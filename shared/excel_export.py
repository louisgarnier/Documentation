"""
Excel Export Module for Test Case Documentation Tool

This module handles exporting test cases and their data to Excel workbooks.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from shared.models import (
    get_all_test_cases,
    get_test_case_by_id,
    get_steps_by_test_case,
    get_screenshots_by_step,
    get_project_by_id,
    get_all_projects
)
from datetime import datetime
import os
from pathlib import Path


def create_excel_export(output_path="test_cases_export.xlsx", selected_test_case_ids=None, selected_project_ids=None):
    """
    Create an Excel workbook with test case documentation.
    
    Args:
        output_path: Path where the Excel file should be saved
        selected_test_case_ids: Optional list of test case IDs to export. If None, exports all.
        selected_project_ids: Optional list of project IDs to export. If provided, exports all test cases from those projects.
        
    Returns:
        str: Path to the created Excel file
    """
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Get test cases based on provided filters
    if selected_project_ids:
        # Fetch test cases by project IDs
        test_cases = []
        projects_dict = {}  # Store project info for grouping
        for project_id in selected_project_ids:
            project = get_project_by_id(project_id)
            if project:
                projects_dict[project_id] = project
                project_test_cases = get_all_test_cases(project_id=project_id)
                # Add project_id to each test case for grouping
                for tc in project_test_cases:
                    tc['_project_id'] = project_id
                    tc['_project_name'] = project['name']
                test_cases.extend(project_test_cases)
        
        # If test_case_ids also provided, filter to intersection
        if selected_test_case_ids:
            test_case_id_set = set(selected_test_case_ids)
            test_cases = [tc for tc in test_cases if tc['id'] in test_case_id_set]
    elif selected_test_case_ids:
        # Fetch specific test cases
        all_test_cases = get_all_test_cases()
        test_cases = [tc for tc in all_test_cases if tc['id'] in selected_test_case_ids]
        # Add project info for grouping
        projects_dict = {}
        for tc in test_cases:
            if tc.get('project_id'):
                project_id = tc['project_id']
                if project_id not in projects_dict:
                    project = get_project_by_id(project_id)
                    if project:
                        projects_dict[project_id] = project
                if project_id in projects_dict:
                    tc['_project_id'] = project_id
                    tc['_project_name'] = projects_dict[project_id]['name']
    else:
        # Export all test cases
        test_cases = get_all_test_cases()
        # Add project info for grouping
        projects_dict = {}
        for tc in test_cases:
            if tc.get('project_id'):
                project_id = tc['project_id']
                if project_id not in projects_dict:
                    project = get_project_by_id(project_id)
                    if project:
                        projects_dict[project_id] = project
                if project_id in projects_dict:
                    tc['_project_id'] = project_id
                    tc['_project_name'] = projects_dict[project_id]['name']
    
    # Group test cases by project for summary
    # Always use projects_dict (populated for all cases)
    projects_info = projects_dict
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    create_summary_sheet(summary_sheet, test_cases, projects_info)
    
    # Create a sheet for each test case
    for test_case in test_cases:
        # Include project name in sheet name if available
        project_name = test_case.get('_project_name', '')
        if project_name:
            # Clean project name for Excel sheet name (remove invalid characters)
            clean_project_name = project_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
            sheet_name = f"{clean_project_name}_{test_case['test_number']}"
        else:
            sheet_name = f"{test_case['test_number']}"
        
        # Excel sheet names have a 31 character limit
        if len(sheet_name) > 31:
            # Truncate intelligently: keep project prefix if possible
            if project_name and len(clean_project_name) < 15:
                # Keep project name and truncate test number
                max_tc_len = 31 - len(clean_project_name) - 1  # -1 for underscore
                tc_part = test_case['test_number'][:max_tc_len]
                sheet_name = f"{clean_project_name}_{tc_part}"
            else:
                sheet_name = sheet_name[:28] + "..."
        
        # Check if sheet name already exists (handle duplicates)
        base_name = sheet_name
        counter = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base_name[:26]}_{counter}"
            counter += 1
        
        test_sheet = wb.create_sheet(sheet_name)
        create_test_case_sheet(test_sheet, test_case)
    
    # Save workbook
    wb.save(output_path)
    return output_path


def create_summary_sheet(sheet, test_cases=None, projects_info=None):
    """Create the summary sheet with all test cases matching the reference format, grouped by project."""
    # Get test cases if not provided
    if test_cases is None:
        test_cases = get_all_test_cases()
    
    # Group test cases by project
    if projects_info is None:
        projects_info = {}
    
    # Group test cases by project_id
    grouped_by_project = {}
    ungrouped_test_cases = []
    
    for tc in test_cases:
        project_id = tc.get('_project_id') or tc.get('project_id')
        if project_id and project_id in projects_info:
            if project_id not in grouped_by_project:
                grouped_by_project[project_id] = []
            grouped_by_project[project_id].append(tc)
        else:
            ungrouped_test_cases.append(tc)
    
    # White fill for all cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Apply white fill to entire sheet (first 100 rows, first 20 columns)
    for row in range(1, 101):
        for col in range(1, 21):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thick black border for the outer box
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    # Add padding: 
    # - Row 2: padding top (bordure supérieure)
    # - Row 3: Title
    # - Row 4: spacing
    # - Row 5: Headers
    # - Row 6+: Data
    # - Last row + 1: padding bottom (bordure inférieure)
    # Columns: A = padding left, B-E = content, F+ = padding right
    
    # Row 3: Title (centered in column C, with padding left from column A)
    title_cell = sheet.cell(row=3, column=3)  # Column C instead of B for left padding
    title_cell.value = "Test Case Documentation"
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = white_fill
    # No border on title cell itself
    
    # Row 5: Headers (starting in column C, with padding left)
    headers = ["Test Case ID", "Test Case Name", "Execution Status", "Outcome"]
    header_font = Font(bold=True, size=11, color="FFFFFF")  # White text, bold
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue background
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 3):  # Start from column C (3) instead of B
        cell = sheet.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.fill = header_fill  # Dark blue background instead of white
    
    # Row 6 onwards: Test case data (with padding from headers), grouped by project
    last_data_row = 5  # Will be updated as we add rows
    current_row = 6
    
    # Separator border (thin border for visual separation, but row is white)
    separator_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Project header fill (slightly darker gray)
    project_header_fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
    project_header_font = Font(bold=True, size=12, color="000000")
    
    # Helper function to get sheet name for a test case (matching create_excel_export logic)
    # Note: This generates the base name; actual sheet creation handles duplicates
    def get_sheet_name_for_test_case(tc):
        project_name = tc.get('_project_name', '')
        if project_name:
            clean_project_name = project_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
            sheet_name = f"{clean_project_name}_{tc['test_number']}"
        else:
            sheet_name = f"{tc['test_number']}"
        
        if len(sheet_name) > 31:
            if project_name and len(clean_project_name) < 15:
                max_tc_len = 31 - len(clean_project_name) - 1
                tc_part = tc['test_number'][:max_tc_len]
                sheet_name = f"{clean_project_name}_{tc_part}"
            else:
                sheet_name = sheet_name[:28] + "..."
        
        return sheet_name
    
    # Helper function to write a test case row
    def write_test_case_row(row_num, test_case):
        # Column C: Test Case ID (with hyperlink to test case sheet)
        test_case_id_cell = sheet.cell(row=row_num, column=3)
        test_case_id_cell.value = test_case['test_number']
        test_case_id_cell.border = thin_border
        test_case_id_cell.fill = white_fill
        
        # Create hyperlink to test case sheet
        sheet_name = get_sheet_name_for_test_case(test_case)
        test_case_id_cell.hyperlink = f"#{sheet_name}!A1"
        test_case_id_cell.font = Font(color="0563C1", underline="single")
        
        # Column D: Test Case Name (Description)
        desc_cell = sheet.cell(row=row_num, column=4)
        desc_cell.value = test_case['description']
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.fill = white_fill
        
        # Column E: Execution Status
        status_cell = sheet.cell(row=row_num, column=5)
        status_cell.value = "Completed"
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.fill = white_fill
        
        # Column F: Outcome
        outcome_cell = sheet.cell(row=row_num, column=6)
        outcome_cell.value = "Pass"
        outcome_cell.border = thin_border
        outcome_cell.alignment = Alignment(horizontal="center")
        outcome_cell.fill = white_fill
    
    # Write grouped projects first
    if grouped_by_project:
        for project_id, project_tcs in grouped_by_project.items():
            project = projects_info.get(project_id)
            project_name = project['name'] if project else f"Project {project_id}"
            
            # Add separator row before project (except for first project)
            if current_row > 6:
                for col in range(2, 8):  # Columns B to G
                    cell = sheet.cell(row=current_row, column=col)
                    cell.fill = white_fill
                    # No border on separator row - just white space
                    cell.border = None
                current_row += 1
            
            # Add project header row
            project_header_cell = sheet.cell(row=current_row, column=3)
            project_header_cell.value = f"Project: {project_name}"
            project_header_cell.font = project_header_font
            project_header_cell.fill = project_header_fill
            project_header_cell.alignment = Alignment(horizontal="left", vertical="center")
            # Merge cells for project header (columns C to F)
            sheet.merge_cells(f'C{current_row}:F{current_row}')
            # Apply borders to merged cells
            for col in range(3, 7):
                cell = sheet.cell(row=current_row, column=col)
                cell.border = thin_border
            # Left and right padding
            for col in [2, 7]:
                cell = sheet.cell(row=current_row, column=col)
                cell.fill = white_fill
                cell.border = thin_border
            current_row += 1
            
            # Write test cases for this project
            for test_case in project_tcs:
                write_test_case_row(current_row, test_case)
                current_row += 1
    
    # Write ungrouped test cases (if any)
    if ungrouped_test_cases:
        # Add separator if we had grouped projects
        if grouped_by_project:
            for col in range(2, 8):
                cell = sheet.cell(row=current_row, column=col)
                cell.fill = white_fill
                # No border on separator row - just white space
                cell.border = None
            current_row += 1
        
        # Add header for ungrouped
        ungrouped_header_cell = sheet.cell(row=current_row, column=3)
        ungrouped_header_cell.value = "Unassigned Test Cases"
        ungrouped_header_cell.font = project_header_font
        ungrouped_header_cell.fill = project_header_fill
        ungrouped_header_cell.alignment = Alignment(horizontal="left", vertical="center")
        sheet.merge_cells(f'C{current_row}:F{current_row}')
        for col in range(3, 7):
            cell = sheet.cell(row=current_row, column=col)
            cell.border = thin_border
        for col in [2, 7]:
            cell = sheet.cell(row=current_row, column=col)
            cell.fill = white_fill
            cell.border = thin_border
        current_row += 1
        
        # Write ungrouped test cases
        for test_case in ungrouped_test_cases:
            write_test_case_row(current_row, test_case)
            current_row += 1
    
    # If no grouping, write all test cases normally
    if not grouped_by_project and not ungrouped_test_cases:
        for test_case in test_cases:
            write_test_case_row(current_row, test_case)
            current_row += 1
    
    last_data_row = current_row - 1
    
    # Apply thick black border around the entire box
    # Box spans from row 2 (padding top) to last_data_row + 1 (padding bottom), columns B to G
    # Add padding: row 2 is top padding, row 3 is title, row 4 is spacing, row 5 is headers, row 6+ is data, last+1 is bottom padding
    # Column B = padding left, Columns C-F = content, Column G = padding right
    start_row = 2  # Top padding row (will have top border)
    end_row = last_data_row + 1  # Bottom padding row (will have bottom border)
    start_col = 2  # Column B (padding left)
    content_end_col = 6    # Column F (last content column - Outcome)
    end_col = 7    # Column G (padding right - where the thick right border goes)
    
    # Top border (row 2 - padding row)
    # Only top and side borders, no internal borders
    # Explicitly set borders for each cell to avoid any unwanted borders
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.fill = white_fill
        # Explicitly set borders: only top and side borders
        if col == start_col:
            # Column B: left and top border
            cell.border = Border(left=thick_border.left, top=thick_border.top)
        elif col == end_col:
            # Column G: right and top border (padding right)
            cell.border = Border(right=thick_border.right, top=thick_border.top)
        else:
            # Middle cells (C, D, E, F): ONLY top border, explicitly no other borders
            cell.border = Border(top=thick_border.top, left=None, right=None, bottom=None)
    
    # Bottom border is handled separately for the padding row below
    # All data rows should have thin borders on all sides
    
    # Left border (column B - padding left) - apply to all rows in the box
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=start_col)
        cell.fill = white_fill
        if row == start_row:
            cell.border = Border(left=thick_border.left, top=thick_border.top)
        elif row == end_row:
            cell.border = Border(left=thick_border.left, bottom=thick_border.bottom)
        else:
            cell.border = Border(left=thick_border.left)
    
    # Right border (column G - padding right) - apply to all rows in the box
    # This is where the thick right border goes, not on column F
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)
        cell.fill = white_fill
        
        # Special handling for row 2 (padding row) - already set above
        if row == start_row:
            # Already handled in the top border section above - skip to avoid overwriting
            continue
            
        # For all other rows, apply right border to column G (padding right)
        if row == end_row:
            cell.border = Border(right=thick_border.right, bottom=thick_border.bottom)
        else:
            cell.border = Border(right=thick_border.right)
    
    # Padding right column (column G) - same logic as column B (left padding)
    # White background, no borders, no content - just padding space
    # Column G IS the padding right (end_col = 7), already has borders set above
    # Just ensure it has white fill and no content for all rows
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)  # Column G
        if cell.value is None:  # Don't overwrite if already set
            cell.value = None  # No content
        # Border and fill already set in the right border section above
    
    # Ensure title row (row 3) and spacing row (row 4) have no internal borders, only side borders
    for row in [3, 4]:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
            if col == start_col:
                # Left border only (column B)
                cell.border = Border(left=thick_border.left)
            elif col == end_col:
                # Right border only (column G - padding right)
                cell.border = Border(right=thick_border.right)
            else:
                # No border for middle cells (C, D, E, F) in title/spacing rows
                cell.border = None
    
    # Ensure bottom padding row (last_data_row + 1) has only bottom and side borders
    bottom_padding_row = last_data_row + 1
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=bottom_padding_row, column=col)
        cell.fill = white_fill
        if col == start_col:
            # Column B: left and bottom border
            cell.border = Border(left=thick_border.left, bottom=thick_border.bottom)
        elif col == end_col:
            # Column G: right and bottom border (padding right)
            cell.border = Border(right=thick_border.right, bottom=thick_border.bottom)
        else:
            # Middle cells (C, D, E, F): only bottom border
            cell.border = Border(bottom=thick_border.bottom)
    
    # Auto-adjust column widths
    sheet.column_dimensions['B'].width = 5  # Padding left
    sheet.column_dimensions['C'].width = 20  # Test Case ID
    sheet.column_dimensions['D'].width = 60  # Test Case Name
    sheet.column_dimensions['E'].width = 18  # Execution Status
    sheet.column_dimensions['F'].width = 15  # Outcome
    sheet.column_dimensions['G'].width = 5  # Padding right (same as left padding)


def create_test_case_sheet(sheet, test_case):
    """Create a sheet for a specific test case matching the reference format with embedded screenshots."""
    # Set white background for entire sheet (first 200 rows, 20 columns)
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row in range(1, 201):
        for col in range(1, 21):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
    
    # Row 2: Title and navigation link on the same row
    # Title in column B
    title_cell = sheet.cell(row=2, column=2)
    title_cell.value = f"{test_case['test_number']} - {test_case['description']}"
    title_cell.font = Font(bold=True, size=12)
    title_cell.fill = white_fill
    
    # Navigation link to the right of title (2-3 columns to the right)
    # "Go to" text - aligned to the right of the cell
    go_to_cell = sheet.cell(row=2, column=5)  # Column E (2 columns to the right of title)
    go_to_cell.value = "Go to"
    go_to_cell.font = Font(size=10)
    go_to_cell.fill = white_fill
    go_to_cell.alignment = Alignment(horizontal="right", vertical="center")
    
    # "[Summary]" link
    summary_link_cell = sheet.cell(row=2, column=6)  # Column F (next to "Go to")
    summary_link_cell.value = "[Summary]"
    summary_link_cell.hyperlink = "#Summary!A1"
    summary_link_cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
    summary_link_cell.fill = white_fill
    
    # Freeze panes: freeze at row 3 (just below "Go to [Summary]" row)
    # This will keep the header row (row 2) visible when scrolling
    sheet.freeze_panes = "A3"
    
    # Get steps
    steps = get_steps_by_test_case(test_case['id'])
    
    if steps:
        # Start steps from row 6 (matching reference format)
        current_row = 6
        
        # Define borders and formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        # Border for step header row: only top and bottom (thick)
        step_header_border = Border(
            left=None,
            right=None,
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
        
        for step in steps:
            # Step description row: entire row (columns B to E or more) with same color and top/bottom borders
            # Format: "1/ Description"
            step_cell = sheet.cell(row=current_row, column=2)
            step_cell.value = f"{step['step_number']}/ {step['description']}"
            step_cell.font = Font(bold=True, size=11)
            step_cell.fill = header_fill
            step_cell.border = step_header_border
            step_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Apply same fill and border to entire row (columns B through E or more)
            for col in range(2, 10):  # Columns B to I
                cell = sheet.cell(row=current_row, column=col)
                cell.fill = header_fill
                cell.border = step_header_border
            
            # Notes row: skip one line, then add notes
            notes_row = current_row + 1
            # Empty row for spacing
            for col in range(2, 10):
                cell = sheet.cell(row=notes_row, column=col)
                cell.fill = white_fill
                cell.value = None
            
            # Notes content on next row
            notes_content_row = notes_row + 1
            # Combine all metadata fields into notes
            notes_parts = []
            if step.get('modules'):
                notes_parts.append(f"Modules: {step['modules']}")
            if step.get('calculation_logic'):
                notes_parts.append(f"Calculation: {step['calculation_logic']}")
            if step.get('configuration'):
                notes_parts.append(f"Configuration: {step['configuration']}")
            notes_text = "\n".join(notes_parts)
            
            if notes_text:
                notes_cell = sheet.cell(row=notes_content_row, column=2)
                notes_cell.value = notes_text
                notes_cell.font = Font(size=10)
                notes_cell.fill = white_fill
                notes_cell.alignment = Alignment(wrap_text=True, vertical="top")
                # Merge cells for notes (columns B to E)
                sheet.merge_cells(f'B{notes_content_row}:E{notes_content_row}')
            
            # Get screenshots for this step
            screenshots = get_screenshots_by_step(step['id'])
            
            # Screenshots row: skip one line after notes, then add screenshots
            if notes_text:
                image_start_row = notes_content_row + 2  # Skip one line after notes
            else:
                image_start_row = notes_row + 1  # If no notes, start after spacing row
            
            # Empty row for spacing before screenshots
            for col in range(2, 10):
                cell = sheet.cell(row=image_start_row, column=col)
                cell.fill = white_fill
                cell.value = None
            
            image_row = image_start_row + 1
            image_col = 2  # Start in column B
            
            if screenshots:
                for idx, screenshot in enumerate(screenshots):
                    screenshot_path = screenshot['file_path']
                    
                    # Check if file exists
                    if os.path.exists(screenshot_path):
                        try:
                            # Load and resize image for better layout
                            img = XLImage(screenshot_path)
                            
                            # Resize image to fit nicely (max width 600px, maintain aspect ratio)
                            max_width = 600
                            if img.width > max_width:
                                ratio = max_width / img.width
                                img.width = int(img.width * ratio)
                                img.height = int(img.height * ratio)
                            
                            # Ensure minimum size for readability
                            min_width = 200
                            if img.width < min_width:
                                ratio = min_width / img.width
                                img.width = int(img.width * ratio)
                                img.height = int(img.height * ratio)
                            
                            # Anchor image to cell (column B) with slight offset for spacing
                            cell_ref = f"{get_column_letter(image_col)}{image_row}"
                            img.anchor = cell_ref
                            
                            # Add image to sheet
                            sheet.add_image(img)
                            
                            # Adjust row height to accommodate image (convert pixels to points: 1 point ≈ 1.33 pixels)
                            # Add extra space for padding
                            row_height = max(int(img.height * 0.75) + 20, 120)
                            sheet.row_dimensions[image_row].height = row_height
                            
                            # Add spacing row after each image (except last)
                            if idx < len(screenshots) - 1:
                                image_row += 1
                                sheet.row_dimensions[image_row].height = 10
                                for col in range(2, 10):
                                    cell = sheet.cell(row=image_row, column=col)
                                    cell.fill = white_fill
                            
                            # Move to next row for next image (stack vertically)
                            image_row += 1
                            
                        except Exception as e:
                            # If image can't be loaded, add text reference
                            error_cell = sheet.cell(row=image_row, column=image_col)
                            error_cell.value = f"[Image: {os.path.basename(screenshot_path)}]"
                            error_cell.font = Font(italic=True, color="808080")
                            error_cell.fill = white_fill
                            image_row += 1
                    else:
                        # File doesn't exist, add text reference
                        missing_cell = sheet.cell(row=image_row, column=image_col)
                        missing_cell.value = f"[Image not found: {os.path.basename(screenshot_path)}]"
                        missing_cell.font = Font(italic=True, color="FF0000")
                        missing_cell.fill = white_fill
                        image_row += 1
                
                # Move to next step (leave space after images)
                current_row = image_row + 2
            else:
                # No screenshots, move to next step after notes
                if notes_text:
                    current_row = notes_content_row + 2
                else:
                    current_row = notes_row + 1
                current_row += 1  # Add one more line spacing before next step
        
        # Auto-adjust column widths
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
    else:
        # No steps
        no_steps_cell = sheet.cell(row=6, column=2)
        no_steps_cell.value = "No steps defined for this test case."
        no_steps_cell.font = Font(italic=True)
        no_steps_cell.fill = white_fill


if __name__ == "__main__":
    # Test the export function
    print("Creating Excel export...")
    output_file = create_excel_export("test_export.xlsx")
    print(f"Excel file created: {output_file}")

