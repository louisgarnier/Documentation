"""
Excel Export Module for Test Case Documentation Tool

This module handles exporting test cases and their data to Excel workbooks.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from models import (
    get_all_test_cases,
    get_test_case_by_id,
    get_steps_by_test_case,
    get_screenshots_by_step
)
from datetime import datetime
import os


def create_excel_export(output_path="test_cases_export.xlsx", selected_test_case_ids=None):
    """
    Create an Excel workbook with test case documentation.
    
    Args:
        output_path: Path where the Excel file should be saved
        selected_test_case_ids: Optional list of test case IDs to export. If None, exports all.
        
    Returns:
        str: Path to the created Excel file
    """
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Get test cases (filtered if IDs provided)
    if selected_test_case_ids:
        all_test_cases = get_all_test_cases()
        test_cases = [tc for tc in all_test_cases if tc['id'] in selected_test_case_ids]
    else:
        test_cases = get_all_test_cases()
    
    # Create Summary sheet
    summary_sheet = wb.create_sheet("Summary", 0)
    create_summary_sheet(summary_sheet, test_cases)
    
    # Create a sheet for each test case
    for test_case in test_cases:
        sheet_name = f"{test_case['test_number']}"
        # Excel sheet names have a 31 character limit
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        
        # Check if sheet name already exists (handle duplicates)
        base_name = sheet_name
        counter = 1
        while sheet_name in [s.title for s in wb.worksheets]:
            sheet_name = f"{base_name[:26]}_{counter}"
            counter += 1
        
        test_sheet = wb.create_sheet(sheet_name)
        create_test_case_sheet(test_sheet, test_case)
    
    # Save workbook
    wb.save(output_path)
    return output_path


def create_summary_sheet(sheet, test_cases=None):
    """Create the summary sheet with all test cases matching the reference format."""
    # Get test cases if not provided
    if test_cases is None:
        test_cases = get_all_test_cases()
    
    # White fill for all cells
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Apply white fill to entire sheet (first 100 rows, first 20 columns)
    for row in range(1, 101):
        for col in range(1, 21):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
    
    # Define borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Thick black border for the outer box
    thick_border = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )
    
    # Add padding: 
    # - Row 2: padding top (bordure supérieure)
    # - Row 3: Title
    # - Row 4: spacing
    # - Row 5: Headers
    # - Row 6+: Data
    # - Last row + 1: padding bottom (bordure inférieure)
    # Columns: A = padding left, B-E = content, F+ = padding right
    
    # Row 3: Title (centered in column C, with padding left from column A)
    title_cell = sheet.cell(row=3, column=3)  # Column C instead of B for left padding
    title_cell.value = "Test Case Documentation"
    title_cell.font = Font(bold=True, size=14)
    title_cell.fill = white_fill
    # No border on title cell itself
    
    # Row 5: Headers (starting in column C, with padding left)
    headers = ["Test Case ID", "Test Case Name", "Execution Status", "Outcome"]
    header_font = Font(bold=True, size=11, color="FFFFFF")  # White text, bold
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue background
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col_num, header in enumerate(headers, 3):  # Start from column C (3) instead of B
        cell = sheet.cell(row=5, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.fill = header_fill  # Dark blue background instead of white
    
    # Row 6 onwards: Test case data (with padding from headers)
    last_data_row = 5  # Will be updated as we add rows
    
    for row_num, test_case in enumerate(test_cases, 6):
        # Ensure all cells in this row have borders (table structure)
        # Column C: Test Case ID (with hyperlink to test case sheet) - shifted to column C
        test_case_id_cell = sheet.cell(row=row_num, column=3)  # Column C instead of B
        test_case_id_cell.value = test_case['test_number']
        test_case_id_cell.border = thin_border
        test_case_id_cell.fill = white_fill
        
        # Create hyperlink to test case sheet
        sheet_name = f"{test_case['test_number']}"
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:28] + "..."
        # Handle duplicate sheet names (same logic as in create_excel_export)
        base_name = sheet_name
        counter = 1
        original_name = sheet_name
        while sheet_name in [s.title for s in sheet.parent.worksheets if s != sheet]:
            sheet_name = f"{base_name[:26]}_{counter}"
            counter += 1
        # Use original name for hyperlink (Excel will handle it)
        test_case_id_cell.hyperlink = f"#{sheet_name}!A1"
        test_case_id_cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
        
        # Column D: Test Case Name (Description) - shifted
        desc_cell = sheet.cell(row=row_num, column=4)  # Column D instead of C
        desc_cell.value = test_case['description']
        desc_cell.border = thin_border
        desc_cell.alignment = Alignment(wrap_text=True, vertical="top")
        desc_cell.fill = white_fill
        
        # Column E: Execution Status (default to "Completed" or leave empty) - shifted
        status_cell = sheet.cell(row=row_num, column=5)  # Column E instead of D
        status_cell.value = "Completed"
        status_cell.border = thin_border
        status_cell.alignment = Alignment(horizontal="center")
        status_cell.fill = white_fill
        
        # Column F: Outcome (default to "Pass" or leave empty) - shifted
        outcome_cell = sheet.cell(row=row_num, column=6)  # Column F instead of E
        outcome_cell.value = "Pass"
        outcome_cell.border = thin_border
        outcome_cell.alignment = Alignment(horizontal="center")
        outcome_cell.fill = white_fill
        
        last_data_row = row_num
    
    # Apply thick black border around the entire box
    # Box spans from row 2 (padding top) to last_data_row + 1 (padding bottom), columns B to G
    # Add padding: row 2 is top padding, row 3 is title, row 4 is spacing, row 5 is headers, row 6+ is data, last+1 is bottom padding
    # Column B = padding left, Columns C-F = content, Column G = padding right
    start_row = 2  # Top padding row (will have top border)
    end_row = last_data_row + 1  # Bottom padding row (will have bottom border)
    start_col = 2  # Column B (padding left)
    content_end_col = 6    # Column F (last content column - Outcome)
    end_col = 7    # Column G (padding right - where the thick right border goes)
    
    # Top border (row 2 - padding row)
    # Only top and side borders, no internal borders
    # Explicitly set borders for each cell to avoid any unwanted borders
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=start_row, column=col)
        cell.fill = white_fill
        # Explicitly set borders: only top and side borders
        if col == start_col:
            # Column B: left and top border
            cell.border = Border(left=thick_border.left, top=thick_border.top)
        elif col == end_col:
            # Column G: right and top border (padding right)
            cell.border = Border(right=thick_border.right, top=thick_border.top)
        else:
            # Middle cells (C, D, E, F): ONLY top border, explicitly no other borders
            cell.border = Border(top=thick_border.top, left=None, right=None, bottom=None)
    
    # Bottom border is handled separately for the padding row below
    # All data rows should have thin borders on all sides
    
    # Left border (column B - padding left) - apply to all rows in the box
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=start_col)
        cell.fill = white_fill
        if row == start_row:
            cell.border = Border(left=thick_border.left, top=thick_border.top)
        elif row == end_row:
            cell.border = Border(left=thick_border.left, bottom=thick_border.bottom)
        else:
            cell.border = Border(left=thick_border.left)
    
    # Right border (column G - padding right) - apply to all rows in the box
    # This is where the thick right border goes, not on column F
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)
        cell.fill = white_fill
        
        # Special handling for row 2 (padding row) - already set above
        if row == start_row:
            # Already handled in the top border section above - skip to avoid overwriting
            continue
            
        # For all other rows, apply right border to column G (padding right)
        if row == end_row:
            cell.border = Border(right=thick_border.right, bottom=thick_border.bottom)
        else:
            cell.border = Border(right=thick_border.right)
    
    # Padding right column (column G) - same logic as column B (left padding)
    # White background, no borders, no content - just padding space
    # Column G IS the padding right (end_col = 7), already has borders set above
    # Just ensure it has white fill and no content for all rows
    for row in range(start_row, end_row + 1):
        cell = sheet.cell(row=row, column=end_col)  # Column G
        if cell.value is None:  # Don't overwrite if already set
            cell.value = None  # No content
        # Border and fill already set in the right border section above
    
    # Ensure title row (row 3) and spacing row (row 4) have no internal borders, only side borders
    for row in [3, 4]:
        for col in range(start_col, end_col + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = white_fill
            if col == start_col:
                # Left border only (column B)
                cell.border = Border(left=thick_border.left)
            elif col == end_col:
                # Right border only (column G - padding right)
                cell.border = Border(right=thick_border.right)
            else:
                # No border for middle cells (C, D, E, F) in title/spacing rows
                cell.border = None
    
    # Ensure bottom padding row (last_data_row + 1) has only bottom and side borders
    bottom_padding_row = last_data_row + 1
    for col in range(start_col, end_col + 1):
        cell = sheet.cell(row=bottom_padding_row, column=col)
        cell.fill = white_fill
        if col == start_col:
            # Column B: left and bottom border
            cell.border = Border(left=thick_border.left, bottom=thick_border.bottom)
        elif col == end_col:
            # Column G: right and bottom border (padding right)
            cell.border = Border(right=thick_border.right, bottom=thick_border.bottom)
        else:
            # Middle cells (C, D, E, F): only bottom border
            cell.border = Border(bottom=thick_border.bottom)
    
    # Auto-adjust column widths
    sheet.column_dimensions['B'].width = 5  # Padding left
    sheet.column_dimensions['C'].width = 20  # Test Case ID
    sheet.column_dimensions['D'].width = 60  # Test Case Name
    sheet.column_dimensions['E'].width = 18  # Execution Status
    sheet.column_dimensions['F'].width = 15  # Outcome
    sheet.column_dimensions['G'].width = 5  # Padding right (same as left padding)


def create_test_case_sheet(sheet, test_case):
    """Create a sheet for a specific test case matching the reference format with embedded screenshots."""
    # Row 2: Title linking to Summary (formula format like reference)
    # Format: =Summary!B6&" - "&Summary!C6
    # For now, we'll use a simple title format
    title_cell = sheet.cell(row=2, column=2)
    title_cell.value = f"{test_case['test_number']} - {test_case['description']}"
    title_cell.font = Font(bold=True, size=12)
    
    # Row 10: Navigation links (matching reference format)
    sheet.cell(row=10, column=10).value = "Go to"
    summary_link_cell = sheet.cell(row=10, column=11)
    summary_link_cell.value = "[Summary]"
    summary_link_cell.hyperlink = "#Summary!A1"
    summary_link_cell.font = Font(color="0563C1", underline="single")  # Blue, underlined
    
    # Get steps
    steps = get_steps_by_test_case(test_case['id'])
    
    if steps:
        # Start steps from row 6 (matching reference format)
        current_row = 6
        
        # Define borders and formatting
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blue
        
        for step in steps:
            # Step description in column B (format: "1  Description")
            step_cell = sheet.cell(row=current_row, column=2)
            step_cell.value = f"{step['step_number']}  {step['description']}"
            step_cell.font = Font(bold=True, size=11)
            step_cell.fill = header_fill
            step_cell.border = thin_border
            step_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Add metadata in adjacent columns if available
            if step.get('modules'):
                modules_cell = sheet.cell(row=current_row, column=3)
                modules_cell.value = f"Modules: {step['modules']}"
                modules_cell.border = thin_border
                modules_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if step.get('calculation_logic'):
                calc_cell = sheet.cell(row=current_row, column=4)
                calc_cell.value = f"Calculation: {step['calculation_logic']}"
                calc_cell.border = thin_border
                calc_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if step.get('configuration'):
                config_cell = sheet.cell(row=current_row, column=5)
                config_cell.value = f"Config: {step['configuration']}"
                config_cell.border = thin_border
                config_cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Get screenshots for this step
            screenshots = get_screenshots_by_step(step['id'])
            
            # Add screenshots below the step description
            if screenshots:
                image_row = current_row + 1
                image_col = 2  # Start in column B
                
                # Add spacing row before screenshots
                sheet.row_dimensions[image_row].height = 10
                image_row += 1
                
                for idx, screenshot in enumerate(screenshots):
                    screenshot_path = screenshot['file_path']
                    
                    # Check if file exists
                    if os.path.exists(screenshot_path):
                        try:
                            # Load and resize image for better layout
                            img = XLImage(screenshot_path)
                            
                            # Resize image to fit nicely (max width 600px, maintain aspect ratio)
                            max_width = 600
                            if img.width > max_width:
                                ratio = max_width / img.width
                                img.width = int(img.width * ratio)
                                img.height = int(img.height * ratio)
                            
                            # Ensure minimum size for readability
                            min_width = 200
                            if img.width < min_width:
                                ratio = min_width / img.width
                                img.width = int(img.width * ratio)
                                img.height = int(img.height * ratio)
                            
                            # Anchor image to cell (column B) with slight offset for spacing
                            cell_ref = f"{get_column_letter(image_col)}{image_row}"
                            img.anchor = cell_ref
                            
                            # Add image to sheet
                            sheet.add_image(img)
                            
                            # Adjust row height to accommodate image (convert pixels to points: 1 point ≈ 1.33 pixels)
                            # Add extra space for padding
                            row_height = max(int(img.height * 0.75) + 20, 120)
                            sheet.row_dimensions[image_row].height = row_height
                            
                            # Add spacing row after each image (except last)
                            if idx < len(screenshots) - 1:
                                image_row += 1
                                sheet.row_dimensions[image_row].height = 10
                            
                            # Move to next row for next image (stack vertically)
                            image_row += 1
                            
                        except Exception as e:
                            # If image can't be loaded, add text reference
                            error_cell = sheet.cell(row=image_row, column=image_col)
                            error_cell.value = f"[Image: {os.path.basename(screenshot_path)}]"
                            error_cell.font = Font(italic=True, color="808080")
                            error_cell.border = thin_border
                            image_row += 1
                    else:
                        # File doesn't exist, add text reference
                        missing_cell = sheet.cell(row=image_row, column=image_col)
                        missing_cell.value = f"[Image not found: {os.path.basename(screenshot_path)}]"
                        missing_cell.font = Font(italic=True, color="FF0000")
                        missing_cell.border = thin_border
                        image_row += 1
                
                # Move to next step (leave space after images)
                current_row = image_row + 2
            else:
                # No screenshots, just move to next step
                current_row += 2
        
        # Auto-adjust column widths
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 30
        sheet.column_dimensions['E'].width = 30
    else:
        # No steps
        sheet.cell(row=6, column=2).value = "No steps defined for this test case."
        sheet.cell(row=6, column=2).font = Font(italic=True)


if __name__ == "__main__":
    # Test the export function
    print("Creating Excel export...")
    output_file = create_excel_export("test_export.xlsx")
    print(f"Excel file created: {output_file}")

